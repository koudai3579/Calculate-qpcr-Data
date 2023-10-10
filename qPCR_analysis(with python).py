#memo
#実行：右上の▷

import errno
import subprocess
from datetime import datetime
from openpyxl.utils import get_column_letter
import openpyxl as excel
from openpyxl.styles import Alignment
import math
import statistics
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.error_bar import ErrorBars
import xlwings as xw


#シート読み込み
sheet = excel.load_workbook("qpcr-data.xlsx", data_only=True).active
#ブック作成
book = excel.Workbook()
resultSheet = book.active

data1 = [] #アクチン
data2 = [] #評価遺伝子

#各列のデータを取得
for i in range(1,7):
    #アクチンの列
    if i == 1:
        for j in range(3,27):
            value = sheet.cell(row=j,column=i).value
            if value != None:
                data1.append(value)
   #評価遺伝子の列           
    if i == 2:
        for j in range(3,27):
            value = sheet.cell(row=j,column=i).value
            if value != None:
                data2.append(value)

#各サンプルの合計数をカウント
length1 = len(data1)
length2 = len(data2)

#入力してもらったサンプル数（N）を取得
n = sheet.cell(row=2,column=4).value
number_of_data = len(data1)/n

#遺伝子名を配列で取得(アクチンと評価遺伝子の2つ)
genes = []
for i in range(1,3):
    gene = sheet.cell(row=2,column=i).value
    if gene == None:
        gene = "未入力"
    genes.append(gene)

resultSheet.append (["計算("+genes[1]+")",genes[0],genes[1],"power","avepower","相対値PR1","相対値平均PR1","S.E.（誤差）"])

#各データの書き込み
# data1(アクチン)
a = 2
for i ,data in enumerate(data1):
    resultSheet["B" + str(a)] = data
    a += 1
# data2(評価遺伝子)
a = 2
for i ,data in enumerate(data2):
    resultSheet["C" + str(a)] = data
    a+=1

#サンプルネームを取得し設定
sampleNames = []
for i in range(5,9):
    sampleName = sheet.cell(row=3,column=i).value
    if sampleName == None:
        sampleName = "未入力"
    sampleNames.append(sampleName)

#サンプル名の書き込みおよびN数に応じたセルの結合
a=2
for number , sampleName in enumerate(sampleNames): 
    if a > length1:
        break
    if a % 1 == 0 or a % n == 0:
        resultSheet.cell(row=a,column=1,value = sampleName)
        resultSheet.merge_cells(range_string="A"+str(a)+":"+"A"+str(n+a-1))
        resultSheet["A"+str(a)].alignment = Alignment(horizontal="center", vertical="center")
    a += n

#計算結果用基礎を書き込み→グラフ作成に利用
resultSheet["K2"] = "相対値平均pr1"
resultSheet["J3"] = sampleNames[0]
resultSheet["J4"] = sampleNames[1]
resultSheet["J5"] = sampleNames[2]
resultSheet["J6"] = sampleNames[3]
resultSheet["J2"] = genes[1]
resultSheet["L2"] = "誤差"

#power値を算出
for i in range(length2):
    target = resultSheet.cell(row=i+2,column=3).value
    housekeeping = resultSheet.cell(row=i+2,column=2).value

    if target == "N/A" or housekeeping == "N/A" :continue

    power = 2 ** (-(target- housekeeping))
    resultSheet.cell(row= (i+2), column=4,value=power)

#avepowerを算出
a=2
for number , sample in enumerate(data2): 
    if a > length2:
        break
    # 計算処理 aからn個下までの列の数値から平均を算出
    if a % 1 == 0 or a % n == 0:
        sum = 0
        copy_n = n
        for i in range(n):
            power_i = resultSheet.cell(row= a+i,column=4).value
            if power_i == None:
                copy_n -= 1
                continue
            else:
                sum += power_i 
            avepower = sum/copy_n
            resultSheet.cell(row=a,column=5,value = avepower)
    a += n

#相対値PR1値を算出
for i in range(length2):
    # power取得
    power = resultSheet.cell(row= (i+2), column=4).value
    if power == None:
        continue
    else:
        #各pr1と比較するbase_avepower取得
        base_avepower = resultSheet.cell(row= 2, column=5).value
        #計算結果を出力
        pr1 = power/base_avepower
        resultSheet.cell(row=i+2,column=6,value = pr1)

#相対値平均PR1算出
c=2
for number , sample in enumerate(data2): 
    if c > length2:
        break
    # 計算処理 aからn個下までの列の数値から平均を算出
    if c % 1 == 0 or c % n == 0:
        sum = 0
        copy_n = n
        for i in range(n):
            pr1 = resultSheet.cell(row= c+i,column=6).value
            if pr1 == None:
                copy_n -= 1
                continue
            else:
                sum += pr1
            avepr1 = sum/copy_n
            resultSheet.cell(row=c,column=7,value = avepr1)
    c += n

#求めた相対値平均PR1を要約スペース(グラフ作成用)にも書き込み
e = 2
for i in range(n-1):
    avepr1 = resultSheet. cell(row=e,column=7).value
    resultSheet.cell(row=i+3,column=11,value=avepr1)
    e += n

#誤算算出
a = 2
for number , sample in enumerate(data2): 
    if a > length2:
        break
    if a % 1 == 0 or a % n == 0:
        pr1_row = []
        copy_n = n
        for i in range(n):
            pr1 = resultSheet.cell(row= a+i,column=6).value

            if pr1 == None:
                copy_n -= 1
                continue
            else:
                pr1_row.append(pr1)

        pr1_stdev = statistics.stdev(pr1_row)
        gosa = pr1_stdev / math.sqrt(copy_n)
        resultSheet.cell(row=a,column=8,value = gosa)
    a += n

#求めた誤差を要約スペース(グラフ作成用)にも書き込み
e = 2
for i in range(n-1):
    gosa = resultSheet.cell(row=e,column=8).value
    resultSheet.cell(row=i+3,column=12,value=gosa)
    e += n

# グラフ表示
chart = BarChart()
chart.width = 20
chart.height = 14
chart.title = "Result"              
chart.y_axis.title = '発現量'
chart.legend.position = 'b' 
chart_data = Reference(resultSheet, min_col=11, max_col=11, min_row=2, max_row=2 + number_of_data)
chart_category = Reference(resultSheet, min_col=10, min_row=3, max_row=2 + number_of_data)
chart.add_data(chart_data, titles_from_data=True)
chart.set_categories(chart_category) 
chart.style = 7
chart.type = "col"
chart.grouping = "standard"
chart.gapwidth = 15

# エラーバーを追加してグラフを完成　なぜかエラーバーが表示されていない
error_bars_data = Reference(resultSheet, min_col=12, max_col=12, min_row=3, max_row=2 + number_of_data)
error_bars = ErrorBars()
error_bars.errorBarType = "custom"
chart.series[0].errorBars = error_bars
resultSheet.add_chart(chart, "J10")

# 罫線設定(行目よりn数ごとに分割)
border = Border(bottom=Side(style='thin', color='000000'))
for number , sample in enumerate(data1):
    if (number + 1)  %  n  == 0:
        for j in range(2,9):
            resultSheet.cell(row= number + 2 ,column= j).border = border
            
#セルの幅調節
for col in resultSheet.columns:
    max_length = 0
    column = col[0].column

    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))

    adjusted_width = (max_length + 2) * 1.2
    resultSheet.column_dimensions[get_column_letter(column)].width = adjusted_width

#ファイル名を付け、シートをローカルに保存
current_datetime = datetime.now()
formatted_datetime = current_datetime.strftime("%Y:%m:%-d:%H:%M")
book.save(f"result_qpcr_{formatted_datetime}.xlsx")

#保存したファイルを自動でオープン(iosのみ有効らしい)
excel_file_path = f"result_qpcr_{formatted_datetime}.xlsx"
subprocess.Popen(['open', excel_file_path])